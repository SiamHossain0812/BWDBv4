import csv
import datetime
import openpyxl
import numpy as np
from django.core.files.storage import FileSystemStorage
from django.shortcuts import render
from .models import SpikeData
from django.http import HttpResponse
import pandas as pd
from io import BytesIO, StringIO  # Use BytesIO for binary streams
from datetime import datetime, timedelta

def is_invalid(val):
    """Helper function to check if a value is invalid (9999, -9999999, empty or None)."""
    return str(val).startswith('9999') or str(val).startswith('-9999999') or val == '' or val is None

def is_abnormal(values, index, threshold=2):
    """
    Helper function to check if a value is abnormal based on its deviation from surrounding data.
    A value is considered abnormal if it deviates by more than a threshold from the surrounding data,
    or if it is 0 (which is considered abnormal).
    """
    # Skip invalid values
    if is_invalid(values[index]):
        return False

    # Skip if there aren't enough previous/next values for comparison
    if index < 12 or index + 12 >= len(values):
        return False

    # Check if the value itself is 0, which is considered abnormal
    if values[index] == 0:
        return True

    # Calculate the rolling average and standard deviation for previous and next 12 valid values
    surrounding_values = [v for v in values[max(0, index-12):index+13] if not is_invalid(v)]
    if len(surrounding_values) < 12:
        return False  # Not enough valid data to judge

    mean = np.mean(surrounding_values)
    std_dev = np.std(surrounding_values)

    # Check if the value deviates from the mean by more than 'threshold' standard deviations
    return abs(values[index] - mean) > threshold * std_dev

def replace_invalid_values(values):
    """
    Replace initial invalid values (starts with 9999, -9999999) with 0,
    then replace abnormal (including 0) values with the average 
    of the previous 12 and next 12 valid values.
    """
    values = np.array(values, dtype=float)  # Ensure it's a numpy array for easy manipulation
    abnormal_count = 0  # Track abnormal values count
    invalid_count = 0  # Track invalid values count

    # Step 1: Replace invalid values (e.g., starts with 9999/-9999999) with 0
    for i, val in enumerate(values):
        if str(val).startswith('9999') or str(val).startswith('-9999999') or is_invalid(val):
            values[i] = 0
            invalid_count += 1  # Increment invalid values count

    # Step 2: Handle abnormal values (replace with surrounding mean)
    for i, val in enumerate(values):
        if is_abnormal(values, i):  # Check if the value is abnormal (0 or deviating)
            abnormal_count += 1  # Increment abnormal count

            # Get the previous and next 12 valid values
            prev_values = [v for v in values[max(0, i-12):i] if not is_invalid(v)]
            next_values = [v for v in values[i+1:i+13] if not is_invalid(v)]
            surrounding_values = prev_values + next_values

            # Calculate mean of surrounding valid values
            if surrounding_values:
                mean = np.mean(surrounding_values)
                values[i] = mean  # Replace abnormal value with the mean

    return values.tolist(), invalid_count, abnormal_count


def spikedata(request):
    # Retrieve stored values from the session
    last_uploaded_file_name = request.session.get('uploaded_file_name', '')
    stored_start_date = request.session.get('start_date', '')
    stored_end_date = request.session.get('end_date', '')
    stored_rate_of_change = request.session.get('rate_of_change', '')
    stored_station_id = request.session.get('station_id', '')  # Store station ID instead of name

    # Initialize formatted start and end date variables
    formatted_start_date = ''
    formatted_end_date = ''
    station_name = ''

    # Check if the form is submitted with a POST request
    if request.method == 'POST':
        # Get start and end date from the form
        start_date = request.POST.get('start_date', '')
        end_date = request.POST.get('end_date', '')
        rate_of_change = request.POST.get('rate_of_change', '')
        station_id = request.POST.get('station_name', '')  # Get station id from form

        # Convert start and end date to DD/MM/YYYY format
        try:
            if start_date:
                formatted_start_date = datetime.strptime(start_date, '%Y-%m-%d').strftime('%d/%m/%Y')
            if end_date:
                formatted_end_date = datetime.strptime(end_date, '%Y-%m-%d').strftime('%d/%m/%Y')

            # Store the formatted dates and other values in session
            request.session['start_date'] = formatted_start_date
            request.session['end_date'] = formatted_end_date
            request.session['rate_of_change'] = rate_of_change
            request.session['station_id'] = station_id  # Store station ID in session

        except ValueError:
            return render(request, 'spikedata.html', {'error': 'Invalid date format. Please use YYYY-MM-DD format.'})

        # Retrieve station name by ID from the database
        if station_id:
            station = StationName.objects.filter(id=station_id).first()  # Retrieve station by ID
            if station:
                station_name = station.station_name  # Correctly access the 'station_name' field

        # Handle file upload if a file is provided
        if 'file_upload' in request.FILES:
            file = request.FILES['file_upload']
            fs = FileSystemStorage()
            filename = fs.save(file.name, file)
            filepath = fs.path(filename)

            # Store the uploaded file name in session
            request.session['uploaded_file_name'] = filename
            all_data = []
            total_invalid = 0
            abnormal_count = 0

            # Handle CSV file processing
            if filename.endswith('.csv'):
                try:
                    with open(filepath, 'r', encoding='utf-8') as csvfile:
                        reader = csv.DictReader(csvfile)
                        for row in reader:
                            dateTime_str = row.get('dateTime', '').strip()
                            value = row.get('value', '').strip()

                            try:
                                value = float(value) if value not in ('-', '', ' ') else None
                            except ValueError:
                                value = None

                            if dateTime_str and value is not None:
                                all_data.append({'dateTime': dateTime_str, 'value': value})

                except UnicodeDecodeError:
                    return render(request, 'spikedata.html', {'error': 'Error decoding CSV file. Please check the file encoding and try again.'})

            # Handle Excel file processing
            elif filename.endswith('.xlsx'):
                try:
                    workbook = openpyxl.load_workbook(filepath)
                    for sheet in workbook.worksheets:
                        for row in sheet.iter_rows(min_row=2, values_only=True):
                            dateTime_val = row[0] if row[0] is not None else ''
                            value = row[1] if row[1] is not None else ''

                            try:
                                value = float(value) if value not in ('-', '', ' ') else None
                            except ValueError:
                                value = None

                            if dateTime_val and value is not None:
                                all_data.append({'dateTime': dateTime_val, 'value': value})

                except Exception as e:
                    return render(request, 'spikedata.html', {'error': f'Error processing Excel file: {str(e)}'})

            else:
                return render(request, 'spikedata.html', {'error': 'Unsupported file format.'})

            # Prepare data for summary
            values = [data['value'] for data in all_data]

            # Replace invalid and abnormal values
            replaced_values, invalid_count, abnormal_count = replace_invalid_values(values)

            # Store the data back into the SpikeData model
            SpikeData.objects.all().delete()

            for i, data in enumerate(all_data):
                SpikeData.objects.create(
                    dateTime=data['dateTime'],
                    value=replaced_values[i]
                )

            # Prepare the summary with the stored values included
            summary = {
                'total_data_points': len(all_data),
                'missing_data_points': total_invalid,
                'invalid_data_points': invalid_count,  # Added
                'abnormal_data_points': abnormal_count,
                'last_uploaded_file_name': filename,
                'stored_start_date': formatted_start_date,  # Added
                'stored_end_date': formatted_end_date,      # Added
                'stored_rate_of_change': rate_of_change,    # Added
                'stored_station_name': station_name         # Added
            }

            # Return the rendered page with the updated summary and session data
            return render(request, 'spikedata.html', {
                'success': True,
                'message': 'File uploaded and data analyzed successfully.',
                'summary': summary,
                'stations': StationName.objects.all(),  # Ensure station names are passed
            })

    # Retrieve all station names from the database
    stations = StationName.objects.all()  # Fetch station names from the StationName model

    # Render the page when the form is first loaded or if there's no POST data
    return render(request, 'spikedata.html', {
        'last_uploaded_file_name': last_uploaded_file_name,
        'summary': {
            'last_uploaded_file_name': last_uploaded_file_name,
            'total_data_points': 0,
            'missing_data_points': 0,
            'invalid_data_points': 0,  # Added
            'abnormal_data_points': 0,
            'stored_start_date': stored_start_date,  # Added
            'stored_end_date': stored_end_date,      # Added
            'stored_rate_of_change': stored_rate_of_change,  # Added
            'stored_station_name': station_name      # Display station name
        },
        'stations': stations,
    })









from datetime import datetime
from django.http import HttpResponse
from .models import SpikeData
import csv

def export_spikedata(request):
    # Get the start and end date from the request (or session)
    start_date_str = request.session.get('start_date', '')
    end_date_str = request.session.get('end_date', '')

    # Print the dates for testing
    print(f"Export Start Date: {start_date_str}")
    print(f"Export End Date: {end_date_str}")

    # If both start and end dates are provided
    if start_date_str and end_date_str:
        # Convert start and end dates to datetime objects (expecting DD/MM/YYYY format)
        start_date = datetime.strptime(start_date_str, '%d/%m/%Y')
        end_date = datetime.strptime(end_date_str, '%d/%m/%Y')

        # Set time to 00:00 for start date and 23:59 for end date to cover full days
        start_date_str = start_date.strftime('%d/%m/%Y') + " 00:00"
        end_date_str = end_date.strftime('%d/%m/%Y') + " 23:59"

        # Filter the data between the specified date range
        spike_data_records = SpikeData.objects.filter(
            dateTime__gte=start_date_str,  # Greater than or equal to start date
            dateTime__lte=end_date_str     # Less than or equal to end date
        )

    elif start_date_str:
        # If only start date is provided
        start_date = datetime.strptime(start_date_str, '%d/%m/%Y')
        start_date_str = start_date.strftime('%d/%m/%Y') + " 00:00"
        spike_data_records = SpikeData.objects.filter(dateTime__gte=start_date_str)

    elif end_date_str:
        # If only end date is provided
        end_date = datetime.strptime(end_date_str, '%d/%m/%Y')
        end_date_str = end_date.strftime('%d/%m/%Y') + " 23:59"
        spike_data_records = SpikeData.objects.filter(dateTime__lte=end_date_str)

    else:
        # If no date range is provided, return all records
        spike_data_records = SpikeData.objects.all()

    # Create the HttpResponse object with the appropriate CSV header.
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="spike_data.csv"'

    writer = csv.writer(response)
    writer.writerow(['dateTime', 'value'])  # Write header

    # Write each record in the filtered data
    for record in spike_data_records:
        # Convert record.dateTime to datetime object
        record_dateTime = datetime.strptime(record.dateTime, '%d/%m/%Y %H:%M')
        # Write the record to the CSV
        writer.writerow([record_dateTime.strftime('%d/%m/%Y %H:%M'), record.value])

    return response



import openpyxl
from django.shortcuts import render
from django.http import HttpResponse
from django.db import connection

def upload_excel(request):
    if request.method == 'POST' and request.FILES['excel_file']:
        excel_file = request.FILES['excel_file']
        
        # Load the Excel file
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active

        # Prepare the raw SQL query to insert the data (without specifying id)
        insert_query = """
        INSERT INTO uiapp_stationname (station_name) VALUES (%s);
        """
        
        # List to collect the data for batch insertion
        data_to_insert = []

        # Loop through the rows in the Excel file and prepare data for insertion
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
            station_name = row[0]  # Assuming station names are in the first column
            if station_name:  # Check if the station name is not empty
                data_to_insert.append((station_name,))

        # Execute the raw SQL query in batches to insert the data
        with connection.cursor() as cursor:
            # Execute the insertions in batches of 2000
            batch_size = 500
            for i in range(0, len(data_to_insert), batch_size):
                cursor.executemany(insert_query, data_to_insert[i:i + batch_size])

        return HttpResponse("<h1>Data successfully uploaded!</h1>")

    return render(request, 'upload_excel.html')


from django.http import JsonResponse
from .models import StationName

def get_stations(request):
    stations = StationName.objects.values('id', 'station_name')
    return JsonResponse(list(stations), safe=False)

